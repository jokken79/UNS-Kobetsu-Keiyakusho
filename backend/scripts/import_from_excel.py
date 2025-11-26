"""
Script de Importación desde Excel
Importa datos desde 個別契約書TEXPERT2025.7.xlsx al sistema web

Uso:
    python scripts/import_from_excel.py --file "path/to/excel.xlsx" --mode employees
    python scripts/import_from_excel.py --file "path/to/excel.xlsx" --mode factories
    python scripts/import_from_excel.py --file "path/to/excel.xlsx" --mode all
"""

import sys
import argparse
from pathlib import Path
from datetime import datetime

# Añadir el directorio raíz al path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from sqlalchemy.orm import Session
from app.core.database import SessionLocal
from app.models.employee import Employee
from app.models.factory import Factory


def import_employees_from_excel(file_path: str, db: Session) -> dict:
    """
    Importa empleados desde la hoja DBGenzai

    Args:
        file_path: Ruta al archivo Excel
        db: Sesión de base de datos

    Returns:
        dict con estadísticas de importación
    """
    print(f"\n📊 Importando empleados desde DBGenzai...")

    wb = openpyxl.load_workbook(file_path, data_only=True)

    if 'DBGenzai' not in wb.sheetnames:
        raise ValueError("No se encontró la hoja 'DBGenzai' en el archivo Excel")

    ws = wb['DBGenzai']

    stats = {
        'total': 0,
        'imported': 0,
        'updated': 0,
        'skipped': 0,
        'errors': []
    }

    # Leer headers para verificar estructura
    headers = [cell.value for cell in list(ws.rows)[0]]
    print(f"📋 Columnas encontradas: {len(headers)}")

    # Saltar header (fila 1), procesar desde fila 2
    for row_num, row in enumerate(list(ws.rows)[1:], start=2):
        stats['total'] += 1

        try:
            # Mapeo de columnas (ajustar según estructura real)
            status_raw = row[0].value  # 現在
            employee_number = row[1].value  # 社員№
            派遣先ID = row[2].value  # 派遣先ID
            派遣先 = row[3].value  # 派遣先
            配属先 = row[4].value  # 配属先
            配属ライン = row[5].value  # 配属ライン
            仕事内容 = row[6].value  # 仕事内容
            full_name = row[7].value  # 氏名
            katakana_name = row[8].value  # カナ
            gender = row[9].value  # 性別
            nationality = row[10].value  # 国籍
            date_of_birth = row[11].value  # 生年月日
            age = row[12].value  # 年齢
            hourly_rate = row[13].value  # 時給

            # Validaciones básicas
            if not employee_number or not full_name:
                stats['skipped'] += 1
                continue

            # Convertir employee_number a string
            employee_number = str(employee_number)

            # Determinar status
            if status_raw == '退社':
                status = 'resigned'
            else:
                status = 'active'

            # Buscar empleado existente
            employee = db.query(Employee).filter_by(
                employee_number=employee_number
            ).first()

            if employee:
                # Actualizar empleado existente
                employee.full_name = full_name
                employee.katakana_name = katakana_name
                employee.gender = gender
                employee.nationality = nationality
                employee.date_of_birth = date_of_birth
                employee.status = status
                employee.hourly_rate = hourly_rate
                stats['updated'] += 1
                action = "actualizado"
            else:
                # Crear nuevo empleado
                employee = Employee(
                    employee_number=employee_number,
                    full_name=full_name,
                    katakana_name=katakana_name,
                    gender=gender,
                    nationality=nationality,
                    date_of_birth=date_of_birth,
                    status=status,
                    hourly_rate=hourly_rate
                )
                db.add(employee)
                stats['imported'] += 1
                action = "importado"

            # Commit cada 100 registros
            if stats['total'] % 100 == 0:
                db.commit()
                print(f"  ✓ {stats['total']} registros procesados...")

        except Exception as e:
            error_msg = f"Error en fila {row_num}: {str(e)}"
            stats['errors'].append(error_msg)
            print(f"  ⚠️  {error_msg}")
            continue

    # Commit final
    db.commit()
    wb.close()

    print(f"\n✅ Importación de empleados completada:")
    print(f"  📊 Total procesado: {stats['total']}")
    print(f"  ➕ Nuevos importados: {stats['imported']}")
    print(f"  🔄 Actualizados: {stats['updated']}")
    print(f"  ⏭️  Omitidos: {stats['skipped']}")
    print(f"  ❌ Errores: {len(stats['errors'])}")

    return stats


def import_factories_from_excel(file_path: str, db: Session) -> dict:
    """
    Importa empresas/fábricas desde la hoja TBKaisha

    Args:
        file_path: Ruta al archivo Excel
        db: Sesión de base de datos

    Returns:
        dict con estadísticas de importación
    """
    print(f"\n🏭 Importando empresas/fábricas desde TBKaisha...")

    wb = openpyxl.load_workbook(file_path, data_only=True)

    if 'TBKaisha' not in wb.sheetnames:
        raise ValueError("No se encontró la hoja 'TBKaisha' en el archivo Excel")

    ws = wb['TBKaisha']

    stats = {
        'total': 0,
        'imported': 0,
        'updated': 0,
        'skipped': 0,
        'errors': []
    }

    # Leer headers
    headers = [cell.value for cell in list(ws.rows)[0]]
    print(f"📋 Columnas encontradas: {len(headers)}")

    # Procesar filas
    for row_num, row in enumerate(list(ws.rows)[1:], start=2):
        stats['total'] += 1

        try:
            # Mapeo de columnas
            company_name = row[0].value  # 派遣先
            company_address = row[1].value  # 派遣先住所
            company_phone = row[2].value  # 派遣先電話
            supervisor_dept = row[3].value  # 派遣先責任者部署
            supervisor_name = row[4].value  # 派遣先責任者名
            supervisor_phone = row[5].value  # 派遣先責任者電話
            factory_name = row[6].value  # 工場名
            factory_address = row[7].value  # 工場住所
            factory_phone = row[8].value  # 工場電話
            department = row[9].value  # 配属先

            # Buscar columna ライン (puede variar)
            # Asumiendo que está en columna 13 (índice 12) o después
            line = None
            for i, header in enumerate(headers):
                if header and 'ライン' in str(header):
                    line = row[i].value
                    break

            if not line and len(row) > 12:
                line = row[12].value  # Intentar columna 13

            # Validaciones básicas
            if not company_name or not factory_name:
                stats['skipped'] += 1
                continue

            # Buscar factory existente
            # La clave única es: company_name + factory_name + department + line
            factory = db.query(Factory).filter_by(
                company_name=company_name,
                factory_name=factory_name,
                department=department,
                line=line
            ).first()

            if factory:
                # Actualizar
                factory.company_address = company_address
                factory.company_phone = company_phone
                factory.factory_address = factory_address
                factory.factory_phone = factory_phone
                factory.supervisor_department = supervisor_dept
                factory.supervisor_name = supervisor_name
                factory.supervisor_phone = supervisor_phone
                stats['updated'] += 1
            else:
                # Crear nueva
                factory = Factory(
                    company_name=company_name,
                    company_address=company_address,
                    company_phone=company_phone,
                    factory_name=factory_name,
                    factory_address=factory_address,
                    factory_phone=factory_phone,
                    department=department,
                    line=line,
                    supervisor_department=supervisor_dept,
                    supervisor_name=supervisor_name,
                    supervisor_phone=supervisor_phone
                )
                db.add(factory)
                stats['imported'] += 1

            # Commit cada 50 registros
            if stats['total'] % 50 == 0:
                db.commit()
                print(f"  ✓ {stats['total']} registros procesados...")

        except Exception as e:
            error_msg = f"Error en fila {row_num}: {str(e)}"
            stats['errors'].append(error_msg)
            print(f"  ⚠️  {error_msg}")
            continue

    # Commit final
    db.commit()
    wb.close()

    print(f"\n✅ Importación de empresas/fábricas completada:")
    print(f"  📊 Total procesado: {stats['total']}")
    print(f"  ➕ Nuevas importadas: {stats['imported']}")
    print(f"  🔄 Actualizadas: {stats['updated']}")
    print(f"  ⏭️  Omitidas: {stats['skipped']}")
    print(f"  ❌ Errores: {len(stats['errors'])}")

    return stats


def main():
    """Función principal del script"""
    parser = argparse.ArgumentParser(
        description='Importar datos desde Excel al sistema Kobetsu Keiyakusho'
    )
    parser.add_argument(
        '--file',
        type=str,
        required=True,
        help='Ruta al archivo Excel'
    )
    parser.add_argument(
        '--mode',
        type=str,
        choices=['employees', 'factories', 'all'],
        default='all',
        help='Modo de importación: employees, factories, o all'
    )

    args = parser.parse_args()

    # Verificar que el archivo existe
    if not Path(args.file).exists():
        print(f"❌ Error: El archivo '{args.file}' no existe")
        sys.exit(1)

    print(f"\n{'='*60}")
    print(f"🚀 IMPORTACIÓN DESDE EXCEL")
    print(f"{'='*60}")
    print(f"📁 Archivo: {args.file}")
    print(f"🎯 Modo: {args.mode}")
    print(f"⏰ Inicio: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    db = SessionLocal()

    try:
        if args.mode in ['employees', 'all']:
            employee_stats = import_employees_from_excel(args.file, db)

        if args.mode in ['factories', 'all']:
            factory_stats = import_factories_from_excel(args.file, db)

        print(f"\n{'='*60}")
        print(f"✅ IMPORTACIÓN COMPLETADA EXITOSAMENTE")
        print(f"{'='*60}")
        print(f"⏰ Fin: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    except Exception as e:
        print(f"\n{'='*60}")
        print(f"❌ ERROR EN IMPORTACIÓN")
        print(f"{'='*60}")
        print(f"Error: {str(e)}")

        import traceback
        traceback.print_exc()

        db.rollback()
        sys.exit(1)

    finally:
        db.close()


if __name__ == "__main__":
    main()
